import argparse
import json
import openpyxl
import glob
import sys
from typing import NamedTuple, List
import datetime


def load_workbook(f):
    stdout = sys.stdout
    sys.stdout = None
    workbook = openpyxl.load_workbook(f)
    sys.stdout = stdout
    return workbook


def get_fields(filename):
    resource_id = filename.split("/")[-2]
    workbook = load_workbook(filename)

    if len(workbook._sheets) != 1:
        return

    (sheet,) = workbook._sheets
    first_row = sheet[1]
    field_names = [c.value for c in first_row]

    if len(first_row) > 15 or len(first_row) < 3:
        return

    # Reject first rows with non strings
    if any((not isinstance(c.value, str) for c in first_row)):
        return

    sheet_res = []
    for row in sheet.iter_rows(min_row=2):
        row_ret = {}
        row_ret["resource_id"] = resource_id
        for cell in row:
            if not cell.value:
                continue
            i = cell.column - 1
            title = first_row[i].value
            val = cell.value
            # if isinstance(val, datetime.datetime) or isinstance(val, datetime.time):
            if val is not None:
                val = str(val)
            row_ret[title] = val
        sheet_res.append(row_ret)
    return sheet_res


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    return parser.parse_args()


def main():
    args = parse_args()
    fields = get_fields(args.filename)
    if fields:
        for field in fields:
            print(json.dumps(field))


main()
